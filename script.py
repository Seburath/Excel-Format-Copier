from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.utils import coordinate_from_string, column_index_from_string

formato = load_workbook(filename = 'formato_entrada.xlsx')
resumen = formato["Resumen"]


def copiar_rango(hoja, inicio_x, final_x, inicio_y, final_y, destino_x, destino_y):

    #se traspasan las dimensiones de 'formato_entrada.xlsx' a 'cuadro de cargas.xlsx'
    for y in range (inicio_y,final_y):
        if resumen.row_dimensions[y].height != None:
            hoja.row_dimensions[destino_y+y].height = resumen.row_dimensions[y].height
    for x in range(inicio_x,final_x):
        if resumen.column_dimensions[resumen.cell(column=x, row=1).column].width != None:
            hoja.column_dimensions[resumen.cell(column=destino_x+x, row=1).column].width = resumen.column_dimensions[resumen.cell(column=x, row=1).column].width

    #barrido para cada celda
    for x in range(inicio_x, final_x):
        for y in range (inicio_y, final_y):

            font, border, fill, al = cargar_atributos(x, y)

            #se asignan a la celda correspondiente en 'cuadro de cargas.xlsx'
            hoja.cell(column=destino_x + x, row=destino_y + y).font = font
            hoja.cell(column=destino_x + x, row=destino_y + y).border = border
            hoja.cell(column=destino_x + x, row=destino_y + y).fill = fill
            hoja.cell(column=destino_x + x, row=destino_y + y).alignment = al
            hoja.cell(column=destino_x + x, row=destino_y + y).value = resumen.cell(column=x, row=y).value


def cargar_atributos(x, y):
#se crea atributos basados en cada celda de 'formato_entrada.xlsx'
    #fuentes
    font = Font(name=resumen.cell(column=x, row=y).font.name,
                    size=resumen.cell(column=x, row=y).font.size,
                    bold=resumen.cell(column=x, row=y).font.bold,
                    italic=resumen.cell(column=x, row=y).font.italic,
                    vertAlign=resumen.cell(column=x, row=y).font.vertAlign,
                    underline=resumen.cell(column=x, row=y).font.strike,
                    strike=resumen.cell(column=x, row=y).font.strike,
                    color=resumen.cell(column=x, row=y).font.color)

    #borde
    border = Border(left=Side(border_style=resumen.cell(column=x, row=y).border.left.border_style,
                                color=resumen.cell(column=x, row=y).border.left.color),
                    right=Side(border_style=resumen.cell(column=x, row=y).border.right.border_style,
                                color=resumen.cell(column=x, row=y).border.right.color),
                    top=Side(border_style=resumen.cell(column=x, row=y).border.top.border_style,
                                color=resumen.cell(column=x, row=y).border.top.color),
                    bottom=Side(border_style=resumen.cell(column=x, row=y).border.bottom.border_style,
                                color=resumen.cell(column=x, row=y).border.bottom.color)
                    )

    #relleno
    fill = PatternFill(patternType =resumen.cell(column=x, row=y).fill.patternType,
                        fill_type = resumen.cell(column=x, row=y).fill.fill_type,
                        fgColor = resumen.cell(column=x, row=y).fill.fgColor,
                        start_color = resumen.cell(column=x, row=y).fill.start_color,
                        bgColor = resumen.cell(column=x, row=y).fill.bgColor,
                        end_color = resumen.cell(column=x, row=y).fill.end_color)

    #alineacion
    al = Alignment(horizontal=resumen.cell(column=x, row=y).alignment.horizontal,
                  vertical=resumen.cell(column=x, row=y).alignment.vertical,
                  textRotation=resumen.cell(column=x, row=y).alignment.textRotation,
                  wrapText=resumen.cell(column=x, row=y).alignment.wrapText,
                  shrinkToFit=resumen.cell(column=x, row=y).alignment.shrinkToFit,
                  indent=resumen.cell(column=x, row=y).alignment.indent,
                  relativeIndent=resumen.cell(column=x, row=y).alignment.relativeIndent,
                  justifyLastLine=resumen.cell(column=x, row=y).alignment.justifyLastLine,
                  readingOrder=resumen.cell(column=x, row=y).alignment.readingOrder,
                  )

    return(font,border, fill, al)

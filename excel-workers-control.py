mport os
import datetime
from openpyxl import load_workbook
from termcolor import colored


def confirm_dates():
    data = load_workbook(filename='data.xlsx')
    mes = data["Marzo 2019"]
    now = datetime.datetime.now()
    for i in range(2, mes.max_row):
        exit_row = read(mes, 5, i)
        if now.strftime("%Y-%m-%d") != read(mes, 1, i):
            if exit_row == None :
                write(mes, 5, i, "no ha salido")
    data.save('data.xlsx')


def read(sheet, col, row):
    return sheet.cell(column=col, row=row).value


def write(sheet, col, row, val):
    sheet.cell(column=col, row=row).value = val
    return


def long_month(str_month):
    now = datetime.datetime.now()
    month = {
        '01': 'Enero',
        '02': 'Febrero',
        '03': 'Marzo',
        '04': 'Abril',
        '05': 'Mayo',
        '06': 'Junio',
        '07': 'Julio',
        '08': 'Agosto',
        '09': 'Septiembre',
        '10': 'Octubre',
        '11': 'Noviembre',
        '12': 'Deciembre'
    }
    return month[str_month] + ' ' + now.strftime("%Y")


def crear_cabeceras(sheet):
    write(sheet, 1, 1, "Fecha")
    write(sheet, 2, 1, "Nombre")
    write(sheet, 3, 1, "Codigo")
    write(sheet, 4, 1, "Ingreso")
    write(sheet, 5, 1, "Salida")
    write(sheet, 6, 1, "Total horas")


def log():
    data = load_workbook(filename='data.xlsx')
    personal = data["Personal"]
    now = datetime.datetime.now()
    os.system('clear')
    actual_sheet = long_month(now.strftime("%m"))
    sheet_list = data.sheetnames

    for i in range(2, personal.max_row):
        if read(personal, 3, i) == 0:
            print(colored (read(personal, 1, i), "red"))
        else:
            print(colored (read(personal, 1, i), "green"))

    input_code = int(input('ingrese codigo: '))
    if actual_sheet not in sheet_list:
        data.create_sheet(actual_sheet)
        crear_cabeceras(data[actual_sheet])
        print('Nueva hoja creada: "' + actual_sheet + '"')

    mes = data[actual_sheet]
    new_row = mes.max_row + 1
    for i in range(2, personal.max_row):
        excel_code = read(personal, 2, i)
        entry_row = read(personal, 3, i)
        if input_code == excel_code:
            if entry_row == 0:
                write(personal, 3, i, new_row)
                write(mes, 1, new_row, (now.strftime("%Y-%m-%d")))
                write(mes, 2, new_row, read(personal, 1, i))
                write(mes, 3, new_row, excel_code)
                write(mes, 4, new_row, (now.strftime("%H:%M")))
            else:
                entry_time = read(mes, 4, entry_row)
                print(mes, 4, entry_row)
                entry_time_date = datetime.datetime.strptime(entry_time,"%H:%M")
                exit_time = now.strftime("%H:%M")
                exit_time_date = datetime.datetime.strptime(exit_time,"%H:%M")
                total_time = str(exit_time_date - entry_time_date)
                total_time_date = datetime.datetime.strptime(total_time,"%H:%M:%S")
                total_time = total_time_date.strftime("%H:%M")
                write(mes, 5, entry_row, exit_time)
                write(mes, 6, entry_row, total_time)
                write(personal, 3, i, 0)
    data.save('data.xlsx')


confirm_dates()

while True:
log()
